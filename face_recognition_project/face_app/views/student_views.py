from django.http import JsonResponse, HttpResponse
from django.template.defaultfilters import last
from ..models import StudentUser, ClassGroup, CourseStudentTeacher, LeaveRequest, StudentAttendance, FaceInfo
from ..utils import success_response, error_response
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.authentication import TokenAuthentication
from rest_framework_simplejwt.authentication import JWTAuthentication
from django.contrib.auth.hashers import make_password
from django.utils import timezone
from django.db import transaction
from django.conf import settings
from ..decorators import custom_auth_required, get_token_from_request, get_token_data
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io
import os



# 学生查询接口
@api_view(['GET'])
@custom_auth_required(roles=['teacher','admin'])
def student_query_api(request):
    all_students = []
    try:
        token_data = get_token_from_request(request)
        user_type = token_data.get('user_type')
        user_id = token_data.get('user_id')
        if user_type == 'admin':
            students = StudentUser.objects.all()
        elif user_type == 'teacher':
            class_ids = list(CourseStudentTeacher.objects.filter(teacher=user_id)
                             .values_list('class_group__class_id', flat=True))
            class_ids_str = [str(cid) for cid in class_ids]
            students = StudentUser.objects.filter(class_id__in=class_ids_str)
        else:
            return error_response(message="无权限查询学生信息", code=403)

        for student in students:
            class_group = ClassGroup.objects.get(class_id=student.class_id) if student.class_id else None
            all_students.append({
                'student_id': student.student_id,
                'student_name': student.student_name,
                'class_id': student.class_id,
                'class_name': class_group.class_name if class_group else None,
                'gender': student.gender,
                'email': student.email,
                'avatar': student.avatar,
            })

        return success_response(data=all_students, message="学生信息查询成功")
    except Exception as e:
        return error_response(message=f"查询失败: {str(e)}", code=500)


# 学生更新接口
@api_view(['PUT'])
@custom_auth_required(roles=['teacher', 'admin'])
def student_update_api(request):
    try:
        token_data = get_token_from_request(request)
        user_type = token_data.get('user_type')
        
        if user_type not in ['teacher', 'admin']:
            return error_response(message="无权限更新学生信息", code=403)

        original_student_id = request.data.get('original_student_id')
        student_id = request.data.get('student_id')
        student_name = request.data.get('student_name')
        gender = request.data.get('gender')
        class_id = request.data.get('class_id')

        if not all([original_student_id, student_id, student_name, gender, class_id]):
            return error_response(message='所有字段都为必填项', code=400)

        # 查找要更新的学生
        try:
            student = StudentUser.objects.get(student_id=original_student_id)
        except StudentUser.DoesNotExist:
            return error_response(message='学生不存在', code=404)

        # 如果学号发生变化，检查新学号是否已存在
        if student_id != original_student_id:
            if StudentUser.objects.filter(student_id=student_id).exists():
                return error_response(message='新学号已存在', code=400)

        # 检查班级是否存在
        try:
            ClassGroup.objects.get(class_id=class_id)
        except ClassGroup.DoesNotExist:
            return error_response(message='班级不存在', code=400)

        # 更新学生信息
        student.student_id = student_id
        student.student_name = student_name
        student.gender = gender
        student.class_id = class_id
        student.update_time = timezone.now()
        student.save()

        return success_response(message='学生信息更新成功')
    except Exception as e:
        return error_response(message=f'学生信息更新失败: {str(e)}', code=500)


# 学生删除接口
@api_view(['DELETE'])
@custom_auth_required(roles=['teacher', 'admin'])
def student_delete_api(request):
    try:
        token_data = get_token_from_request(request)
        user_type = token_data.get('user_type')
        
        if user_type not in ['teacher', 'admin']:
            return error_response(message="无权限删除学生", code=403)

        student_id = request.data.get('student_id')

        if not student_id:
            return error_response(message='学号为必填项', code=400)

        # 查找要删除的学生
        try:
            student = StudentUser.objects.get(student_id=student_id)
        except StudentUser.DoesNotExist:
            return error_response(message='学生不存在', code=404)

        # 检查关联数据，确保可以安全删除
        validation_errors = []

        # 检查是否有请假记录
        leave_requests = LeaveRequest.objects.filter(student=student)
        if leave_requests.exists():
            validation_errors.append(f'该学生存在 {leave_requests.count()} 条请假记录')

        # 检查是否有考勤记录
        attendance_records = StudentAttendance.objects.filter(student_users_id=student)
        if attendance_records.exists():
            validation_errors.append(f'该学生存在 {attendance_records.count()} 条考勤记录')

        # 如果存在关联数据，返回错误信息
        if validation_errors:
            error_message = '无法删除学生，存在以下关联数据：\n' + '\n'.join(validation_errors)
            return error_response(message=error_message, code=400)

        # 使用事务确保数据一致性
        with transaction.atomic():
            # 删除学生对应的人脸信息和图片文件
            face_info_list = FaceInfo.objects.filter(student_id=student)
            face_count = face_info_list.count()
            if face_count > 0:
                # 先删除图片文件
                for face_info in face_info_list:
                    if face_info.face_img:
                        # 构建图片文件的完整路径
                        image_path = os.path.join(settings.MEDIA_ROOT, face_info.face_img.name)
                        # 如果文件存在，则删除
                        if os.path.exists(image_path):
                            try:
                                os.remove(image_path)
                            except OSError as e:
                                # 记录错误但不中断删除流程
                                print(f"删除人脸图片文件失败: {image_path}, 错误: {e}")
                
                # 删除数据库记录
                face_info_list.delete()
            
            # 删除学生
            student.delete()

        return success_response(message='学生删除成功')
    except Exception as e:
        return error_response(message=f'学生删除失败: {str(e)}', code=500)


# 学生新增接口
@api_view(['POST'])
def student_add_api(request):
    try:
        student_id = request.data.get('student_id')
        student_name = request.data.get('student_name')
        class_id = request.data.get('class_id')
        gender = request.data.get('gender')
        
        if StudentUser.objects.filter(student_id=student_id).exists():
            return error_response(message='学生学号已存在', code=400)

        # 创建 StudentUser 对象并关联 User
        StudentUser.objects.create(
            student_id=student_id,
            student_name=student_name,
            gender=gender,
            class_id=class_id,
            password=make_password('hbgc'+student_id[-4:]), 
            create_time=timezone.now(),
            update_time=timezone.now()
        )
        return success_response(message='学生新增成功')
    except Exception as e:
        return error_response(message=f'学生新增失败: {str(e)}', code=500)

# 批量导入学生接口
@api_view(['POST'])
@custom_auth_required(roles=['teacher', 'admin'])
def student_batch_add_api(request):
    try:
        students_data = request.data.get('students', [])
        
        if not students_data:
            return error_response(message='没有提供学生数据', code=400)
        
        error_list = []
        students_to_create = []
        
        # 第一步：验证所有数据，不进行任何数据库操作
        for index, student_data in enumerate(students_data):
            student_id = student_data.get('student_id', '').strip()
            student_name = student_data.get('student_name', '').strip()
            gender = student_data.get('gender', '').strip()
            class_name = student_data.get('class_name', '').strip()
            
            # 验证必填字段
            if not all([student_id, student_name, gender, class_name]):
                error_list.append(f"第{index+1}行：缺少必填字段")
                continue
            
            # 检查学号是否已存在
            if StudentUser.objects.filter(student_id=student_id).exists():
                error_list.append(f"第{index+1}行：学号 {student_id} 已存在")
                continue
            
            # 检查当前批次中是否有重复学号
            existing_ids = [s['student_id'] for s in students_to_create]
            if student_id in existing_ids:
                error_list.append(f"第{index+1}行：学号 {student_id} 在导入数据中重复")
                continue
            
            # 查找班级ID
            try:
                class_group = ClassGroup.objects.get(class_name=class_name)
                class_id = class_group.class_id
            except ClassGroup.DoesNotExist:
                error_list.append(f"第{index+1}行：班级 {class_name} 不存在")
                continue
            
            # 验证性别
            if gender not in ['男', '女']:
                error_list.append(f"第{index+1}行：性别只能是'男'或'女'")
                continue
            
            # 如果所有验证都通过，添加到待创建列表
            students_to_create.append({
                'student_id': student_id,
                'student_name': student_name,
                'gender': gender,
                'class_id': class_id
            })
        
        # 如果有任何错误，直接返回错误，不进行任何数据库操作
        if error_list:
            return error_response(
                message='数据验证失败，请修正以下错误后重新导入',
                data={'errors': error_list},
                code=400
            )
        
        # 第二步：使用事务批量创建所有学生
        try:
            with transaction.atomic():
                created_students = []
                for student_data in students_to_create:
                    student = StudentUser.objects.create(
                        student_id=student_data['student_id'],
                        student_name=student_data['student_name'],
                        gender=student_data['gender'],
                        class_id=student_data['class_id'],
                        password=make_password('hbgc' + student_data['student_id'][-4:]),
                        create_time=timezone.now(),
                        update_time=timezone.now()
                    )
                    created_students.append(student)
                
                return success_response(
                    data={'success_count': len(created_students)},
                    message=f'批量导入成功，共导入 {len(created_students)} 名学生'
                )
                
        except Exception as e:
            return error_response(
                message=f'批量导入失败，所有数据已回滚: {str(e)}',
                code=500
            )
            
    except Exception as e:
        return error_response(message=f'批量导入失败: {str(e)}', code=500)

# Excel模板下载接口
@api_view(['GET'])
@custom_auth_required(roles=['teacher', 'admin'])
def student_template_download_api(request):
    try:
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "学生信息"
        
        # 设置表头
        headers = ['学号', '姓名', '性别', '班级']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 添加示例数据
        example_data = [
            ['2024001', '张三', '男', '计算机科学与技术1班'],
            ['2024002', '李四', '女', '软件工程1班'],
            ['2024003', '王五', '男', '网络工程1班']
        ]
        
        for row_idx, row_data in enumerate(example_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 设置列宽
        column_widths = [15, 10, 8, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 创建HTTP响应
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="学生信息导入模板.xlsx"'
        
        return response
        
    except Exception as e:
        return error_response(message=f"模板下载失败: {str(e)}", code=500)

# 获取所有班级列表接口
@api_view(['GET'])
@custom_auth_required(roles=['teacher', 'admin'])
def class_list_api(request):
    try:
        classes = ClassGroup.objects.all().values('class_id', 'class_name')
        class_list = [{'class_id': cls['class_id'], 'class_name': cls['class_name']} for cls in classes]
        
        return success_response(
            data={'classes': class_list},
            message='获取班级列表成功'
        )
        
    except Exception as e:
        return error_response(message=f"获取班级列表失败: {str(e)}", code=500)

# 获取当前登录的学生信息
@api_view(['GET'])
@custom_auth_required(roles=['student'])
def student_selfUser_api(request):
    try:
        user_data = get_token_data(request)
        student = StudentUser.objects.filter(student_users_id=user_data[0]).first()
        if not student:
            return error_response(message="学生不存在", code=404)
        return success_response(data={
            'student_id': student.student_id,
            'student_name': student.student_name
        }, message="学生信息查询成功")
    except Exception as e:
        return error_response(message=f"查询失败: {str(e)}", code=500)